import openpyxl

workbook = openpyxl.load_workbook("../excel_tables/data_base_update.xlsx")
worksheet = workbook['Products']
worksheet_attributes = workbook['ProductAttributes']
worksheet_imgs = workbook['AdditionalImages']

def write(products):
    max_row = worksheet.max_row
    max_row_imgs = worksheet_imgs.max_row
    max_row_attributes = worksheet_attributes.max_row
    for product in products:
        found = False
        sort_order_img = 1
        for i, cell in enumerate(worksheet['L'][1:], start=2):
            if product['model'] == cell.value:
                worksheet[f'P{i}'] = product['price']
                worksheet[f'AL{i}'] = product['sort_order']
                found = True
                break
        if product['main_img'] is None:
            continue
        elif found == False:
            max_row += 1
            worksheet[f'A{max_row}'] = product['id']
            worksheet[f'B{max_row}'] = product['name']
            worksheet[f'C{max_row}'] = product['categories']
            worksheet[f'K{max_row}'] = '0'
            worksheet[f'L{max_row}'] = product['model']
            worksheet[f'M{max_row}'] = product['manufactor']
            worksheet[f'N{max_row}'] = product['main_img']
            worksheet[f'O{max_row}'] = 'yes'
            worksheet[f'P{max_row}'] = f'{product['price']}'
            worksheet[f'Q{max_row}'] = '0'
            worksheet[f'U{max_row}'] = '0'
            worksheet[f'V{max_row}'] = 'кг'
            worksheet[f'W{max_row}'] = '0'
            worksheet[f'X{max_row}'] = '0'
            worksheet[f'Y{max_row}'] = '0'
            worksheet[f'Z{max_row}'] = 'см'
            worksheet[f'AA{max_row}'] = 'true'
            worksheet[f'AB{max_row}'] = '0'
            worksheet[f'AC{max_row}'] = product['description']
            worksheet[f'AD{max_row}'] = product['meta_title']
            worksheet[f'AE{max_row}'] = product['meta_description']
            worksheet[f'AF{max_row}'] = product['meta_keywords']
            worksheet[f'AG{max_row}'] = '7'
            worksheet[f'AH{max_row}'] = '0'
            worksheet[f'AK{max_row}'] = product['tags']
            worksheet[f'AL{max_row}'] = product['sort_order']
            worksheet[f'AM{max_row}'] = 'false'
            worksheet[f'AN{max_row}'] = '1'
            for img in product['additiobal_imgs']:
                max_row_imgs += 1
                worksheet_imgs[f'A{max_row_imgs}'] = product['id']
                worksheet_imgs[f'B{max_row_imgs}'] = img
                worksheet_imgs[f'C{max_row_imgs}'] = str(sort_order_img)
                sort_order_img += 1
            for attribute in product['attributes']:
                max_row_attributes += 1
                worksheet_attributes[f'A{max_row_attributes}'] = product['id']
                worksheet_attributes[f'B{max_row_attributes}'] = attribute['attribute_group']
                worksheet_attributes[f'C{max_row_attributes}'] = attribute['attribute']
                worksheet_attributes[f'D{max_row_attributes}'] = attribute['value']

    workbook.save('../excel_tables/data_base_update.xlsx')